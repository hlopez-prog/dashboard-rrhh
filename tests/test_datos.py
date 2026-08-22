"""
Pruebas de la base de datos en Excel y del validador.

    python3 -m unittest discover -s tests -v

Verifican tres cosas que importan en la operación mensual:
  1. Que el libro se lea igual que se escribió (ida y vuelta sin pérdida).
  2. Que un error de captura realista sea detectado, con la fila correcta.
  3. Que el libro traiga las ayudas de captura que RRHH necesita.
"""
import os
import shutil
import sys
import tempfile
import unittest

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "etl"))

import schema  # noqa: E402
import excel  # noqa: E402
import validate  # noqa: E402
import cargar as cargador  # noqa: E402

from openpyxl import load_workbook  # noqa: E402


def libro_existe():
    return os.path.exists(excel.LIBRO)


@unittest.skipUnless(libro_existe(), "falta data/BASE_RRHH.xlsx")
class TestLecturaLibro(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tablas, cls.config = excel.leer()

    def test_estan_todas_las_hojas(self):
        for tabla in schema.ORDEN_CARGA:
            self.assertIn(tabla, self.tablas, f"falta la hoja {tabla}")

    def test_las_hojas_imprescindibles_traen_registros(self):
        """
        Sin catálogos, plantilla y metas no hay tablero. Una hoja de hechos
        vacía, en cambio, es un módulo que todavía no se captura: el tablero
        lo publica como "—". Exigir aquí que todas traigan filas obligaría a
        inventar un mes de nómina para poder desplegar.
        """
        for tabla in ("dim_unidad", "dim_area", "fact_plantilla", "metas"):
            self.assertGreater(len(self.tablas[tabla]), 0,
                               f"la hoja {tabla} no tiene registros")

    def test_encabezados_exactos_y_en_orden(self):
        wb = load_workbook(excel.LIBRO, read_only=True)
        for tabla in schema.ORDEN_CARGA:
            esperados = list(schema.TABLAS[tabla]["columnas"].keys())
            ws = wb[tabla]
            reales = [ws.cell(row=1, column=j).value
                      for j in range(1, len(esperados) + 1)]
            self.assertEqual(reales, esperados, f"encabezados de {tabla}")
        wb.close()

    def test_los_periodos_se_leen_como_texto_aaaa_mm(self):
        # Excel convierte "2026-08" en fecha si la columna no es texto.
        for fila in self.tablas["fact_plantilla"][:50]:
            self.assertRegex(str(fila["periodo"]), r"^\d{4}-(0[1-9]|1[0-2])$")

    def test_los_numeros_se_leen_como_numeros(self):
        fila = self.tablas["fact_plantilla"][0]
        self.assertIsInstance(fila["headcount"], (int, float))
        self.assertGreater(fila["headcount"], 0)
        if self.tablas["fact_nomina"]:
            costo = self.tablas["fact_nomina"][0]["costo_ordinario"]
            self.assertIsInstance(costo, (int, float))
            self.assertGreater(costo, 0)

    def test_la_configuracion_viene_de_la_hoja_leeme(self):
        self.assertIn(self.config["origen"], ("DEMO", "REAL"))
        self.assertTrue(self.config["organizacion"])

    def test_el_libro_pasa_la_validacion(self):
        errores, _ = validate.validar(excel.leer()[0], "excel")
        self.assertEqual(errores, [], f"primeros errores: {errores[:5]}")

    def test_la_carga_da_precedencia_al_libro_de_excel(self):
        _, _, fuente = cargador.cargar()
        self.assertEqual(fuente, "excel")


@unittest.skipUnless(libro_existe(), "falta data/BASE_RRHH.xlsx")
class TestAyudasDeCaptura(unittest.TestCase):
    """El libro es la herramienta de captura: debe guiar a quien lo llena."""

    @classmethod
    def setUpClass(cls):
        cls.wb = load_workbook(excel.LIBRO)

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()

    def test_tiene_hoja_de_instrucciones_al_inicio(self):
        self.assertEqual(self.wb.sheetnames[0], excel.HOJA_LEEME)

    def test_los_encabezados_son_negros_con_texto_blanco(self):
        ws = self.wb["fact_plantilla"]
        celda = ws["A1"]
        self.assertEqual(celda.fill.fgColor.rgb, excel.NEGRO)
        self.assertEqual(celda.font.color.rgb, excel.BLANCO)
        self.assertTrue(celda.font.bold)

    def test_el_encabezado_queda_congelado(self):
        for tabla in schema.ORDEN_CARGA:
            self.assertEqual(self.wb[tabla].freeze_panes, "A2", f"hoja {tabla}")

    def test_las_columnas_de_catalogo_tienen_desplegable(self):
        ws = self.wb["fact_plantilla"]
        rangos = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
        self.assertTrue(rangos, "la hoja no tiene ninguna validación")
        listas = [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]
        self.assertGreaterEqual(len(listas), 3,
                                "faltan listas para tipo_relacion, turno, unidad_id, area_id")

    def test_la_llave_primaria_esta_anotada_en_la_hoja(self):
        for tabla in schema.ORDEN_CARGA:
            comentario = self.wb[tabla]["A1"].comment
            self.assertIsNotNone(comentario, f"hoja {tabla} sin nota de llave")
            for col in schema.TABLAS[tabla]["pk"]:
                self.assertIn(col, comentario.text, f"hoja {tabla}")

    def test_la_celda_de_origen_solo_acepta_demo_o_real(self):
        ws = self.wb[excel.HOJA_LEEME]
        listas = [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]
        self.assertTrue(any("DEMO" in (dv.formula1 or "") for dv in listas))


@unittest.skipUnless(libro_existe(), "falta data/BASE_RRHH.xlsx")
class TestValidadorAtrapaErrores(unittest.TestCase):
    """
    Errores de captura realistas. Cada uno debe detenerse ANTES de publicar,
    y el mensaje debe traer el número de fila del libro.
    """

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.copia = os.path.join(self.dir, "BASE.xlsx")
        shutil.copy(excel.LIBRO, self.copia)

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    # Valores que no pueden ser un 100 cualquiera: los que participan en una
    # regla de negocio, o en un catálogo cerrado.
    SEMILLA = {
        "tipo_relacion": "Propio", "turno": "7x7", "direccion": "menor_mejor",
        "sindicato": "Sindicato de prueba", "riesgo_sindical": 3, "enps": 0,
        "horas_programadas": 1000, "horas_ausencia": 10,
        "dc3_requeridos": 10, "dc3_emitidos": 5,
        "puestos_criticos_totales": 10, "puestos_criticos_cubiertos": 5,
        "es_corporativo": 0, "es_critica": 1,
    }

    def _sembrar(self, wb, hoja):
        """
        Deja una fila válida en la hoja si está vacía.

        Los tests del validador prueban que un error de captura se detecte;
        no deben depender de qué módulos ya se estén capturando. Sin esto, el
        día que nómina se publique vacía —porque todavía no se captura— estas
        pruebas fallarían sin que nada esté mal.
        """
        ws = wb[hoja]
        if any(c.value not in (None, "") for c in ws[2]):
            return
        cols = list(schema.TABLAS[hoja]["columnas"].items())
        primera = {"unidad_id": wb["dim_unidad"]["A2"].value,
                   "area_id": wb["dim_area"]["A2"].value,
                   "periodo": wb["fact_plantilla"]["A2"].value or "2026-08"}
        for j, (col, tipo) in enumerate(cols, start=1):
            if col in primera:
                v = primera[col]
            elif col in self.SEMILLA:
                v = self.SEMILLA[col]
            elif tipo == "p":
                v = primera["periodo"]
            elif tipo == "s":
                v = f"prueba {col}"
            else:
                v = 100
            ws.cell(row=2, column=j, value=v)

    def _validar_tras_editar(self, hoja, celda, valor):
        wb = load_workbook(self.copia)
        self._sembrar(wb, hoja)
        wb[hoja][celda] = valor
        wb.save(self.copia)
        wb.close()
        tablas, _ = excel.leer(self.copia)
        return validate.validar(tablas, "excel")[0]

    def test_atrapa_ausentismo_mayor_que_horas_programadas(self):
        # horas_ausencia (col E) por arriba de horas_programadas (col D)
        errores = self._validar_tras_editar("fact_ausentismo", "E2", 999999)
        self.assertTrue(any("horas_ausencia <= horas_programadas" in e for e in errores),
                        f"errores obtenidos: {errores[:3]}")
        self.assertTrue(any("fila 2" in e for e in errores))

    def test_atrapa_periodo_mal_escrito(self):
        errores = self._validar_tras_editar("fact_plantilla", "A2", "ago-2026")
        self.assertTrue(any("periodo inválido" in e for e in errores),
                        f"errores obtenidos: {errores[:3]}")

    def test_atrapa_unidad_inexistente(self):
        errores = self._validar_tras_editar("fact_nomina", "B2", "U99")
        self.assertTrue(any("no existe en dim_unidad" in e for e in errores),
                        f"errores obtenidos: {errores[:3]}")

    def test_atrapa_texto_donde_va_numero(self):
        errores = self._validar_tras_editar("fact_plantilla", "F2", "cuarenta")
        self.assertTrue(any("donde se" in e and "número" in e for e in errores),
                        f"errores obtenidos: {errores[:3]}")

    def test_atrapa_valor_fuera_de_catalogo(self):
        errores = self._validar_tras_editar("fact_plantilla", "D2", "Eventual")
        self.assertTrue(any("fuera de catálogo" in e for e in errores),
                        f"errores obtenidos: {errores[:3]}")

    def test_atrapa_dc3_emitidos_mayor_que_requeridos(self):
        errores = self._validar_tras_editar("fact_capacitacion", "H2", 999999)
        self.assertTrue(any("dc3_emitidos <= dc3_requeridos" in e for e in errores),
                        f"errores obtenidos: {errores[:3]}")

    def test_atrapa_llave_duplicada(self):
        wb = load_workbook(self.copia)
        ws = wb["dim_unidad"]
        ws["A3"] = ws["A2"].value  # dos unidades con la misma clave
        wb.save(self.copia)
        wb.close()
        tablas, _ = excel.leer(self.copia)
        errores = validate.validar(tablas, "excel")[0]
        self.assertTrue(any("llave duplicada" in e for e in errores),
                        f"errores obtenidos: {errores[:3]}")

    def test_un_libro_valido_no_produce_errores(self):
        tablas, _ = excel.leer(self.copia)
        self.assertEqual(validate.validar(tablas, "excel")[0], [])


class TestColumnasOpcionales(unittest.TestCase):
    """
    Una columna opcional vacía no es un error, y tampoco es un cero: queda
    en None. Es la diferencia entre "no hubo mujeres en esa área" y "nadie
    capturó cuántas mujeres hubo".
    """

    def _fila_plantilla(self, **cambios):
        base = {
            "periodo": "2026-08", "unidad_id": "U01", "area_id": "A01",
            "tipo_relacion": "Propio", "turno": "7x7", "headcount": 100,
            "dotacion_autorizada": "", "mujeres": "", "antiguedad_prom_meses": "",
            "puestos_criticos_totales": "", "puestos_criticos_cubiertos": "",
            "_n": 2,
        }
        base.update(cambios)
        return base

    def _validar(self, fila):
        tablas = {
            "dim_unidad": [{"unidad_id": "U01", "unidad": "U", "estado": "Zac",
                            "tipo_operacion": "Subterránea",
                            "mineral_principal": "Plata", "es_corporativo": 0,
                            "_n": 2}],
            "dim_area": [{"area_id": "A01", "area": "Mina", "tipo_area": "Mina",
                          "es_critica": 1, "_n": 2}],
            "fact_plantilla": [fila],
        }
        return validate.validar(tablas, "excel")[0], tablas

    def test_opcional_vacia_no_es_error_y_queda_en_none(self):
        errores, tablas = self._validar(self._fila_plantilla())
        self.assertEqual(errores, [], f"errores: {errores[:3]}")
        fila = tablas["fact_plantilla"][0]
        for col in ("dotacion_autorizada", "mujeres", "antiguedad_prom_meses"):
            self.assertIsNone(fila[col], f"{col} debería quedar en None, no en 0")

    def test_obligatoria_vacia_si_es_error(self):
        errores, _ = self._validar(self._fila_plantilla(headcount=""))
        self.assertTrue(any("headcount" in e for e in errores))
        self.assertTrue(any(getattr(e, "clase", "") == "vacia_obligatoria"
                            for e in errores),
                        "el error debe venir clasificado para el diagnóstico")

    def test_turno_vacio_no_dispara_la_regla_de_catalogo(self):
        """Un turno vacío ya se reporta una vez (o ninguna, si es opcional):
        la regla de catálogo no debe reportarlo otra vez."""
        errores, _ = self._validar(self._fila_plantilla(turno=""))
        self.assertEqual([e for e in errores if "catálogo" in e], [],
                         f"errores: {errores[:3]}")


class TestDiagnostico(unittest.TestCase):
    """El diagnóstico agrupa: 400 celdas iguales son un pendiente, no 400."""

    def setUp(self):
        sys.path.insert(0, os.path.join(RAIZ, "etl"))
        import diagnostico  # noqa: PLC0415
        self.d = diagnostico

    def test_rangos_comprime_filas_consecutivas(self):
        self.assertEqual(self.d.rangos([4, 5, 6, 9, 10, 20]), "4-6, 9-10, 20")
        self.assertEqual(self.d.rangos([7]), "7")
        self.assertEqual(self.d.rangos([]), "—")
        self.assertEqual(self.d.rangos([3, None, 2]), "2-3")

    def test_recortar_no_corta_un_numero_a_la_mitad(self):
        texto = ", ".join(str(i) for i in range(1, 41))
        salida = self.d.recortar(texto, piezas=5)
        self.assertTrue(salida.startswith("1, 2, 3, 4, 5"))
        self.assertIn("35 rangos más", salida)

    def test_agrupa_por_hoja_columna_y_clase(self):
        errores = [
            validate.Hallazgo("x", "fact_plantilla", "vacia_obligatoria",
                              "headcount", n)
            for n in range(2, 12)
        ] + [validate.Hallazgo("y", "fact_nomina", "llave_duplicada", "pk", 5)]
        grupos = self.d.agrupar(errores)
        self.assertEqual(len(grupos), 2, "diez celdas iguales son un pendiente")
        (tabla, clase, col), g = grupos[0]
        self.assertEqual((tabla, clase, col),
                         ("fact_plantilla", "vacia_obligatoria", "headcount"))
        self.assertEqual(len(g["filas"]), 10)
        self.assertIn("2-11", self.d.rangos(g["filas"]))

    def test_explica_cada_clase_de_problema(self):
        """Ningún grupo puede quedarse sin instrucción de qué hacer."""
        for clase in self.d.ORDEN_CLASES:
            self.assertIn(clase, self.d.QUE_HACER,
                          f"falta el 'qué hacer' de la clase {clase}")

    def test_detecta_bloques_con_area_repetida(self):
        tablas = {
            "dim_area": [{"area_id": "A01"}, {"area_id": "A02"}],
            "fact_nomina": [
                {"periodo": "2026-08", "unidad_id": "U01", "area_id": "A01"},
                {"periodo": "2026-08", "unidad_id": "U01", "area_id": "A02"},
                {"periodo": "2026-08", "unidad_id": "U01", "area_id": "A02"},
            ],
        }
        b = [x for x in self.d.bloques(tablas) if x["hoja"] == "fact_nomina"][0]
        self.assertEqual(b["formas"], [(3, 1)])
        self.assertEqual(b["repetidas"], [("A02", 1)])
        self.assertEqual(b["areas_catalogo"], 2)


if __name__ == "__main__":
    unittest.main(verbosity=2)
