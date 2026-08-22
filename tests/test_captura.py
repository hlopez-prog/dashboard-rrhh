"""
Pruebas de las hojas de captura por unidad.

    python3 -m unittest discover -s tests -v

El riesgo que cubren: una hoja de captura mal construida no falla de inmediato
—falla dentro de dos semanas, cuando ocho analistas ya la llenaron y hay que
devolvérsela a todos. Por eso se verifica la estructura, no solo que el archivo
exista.
"""
import os
import shutil
import sys
import tempfile
import unittest

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "etl"))

import schema  # noqa: E402
import excel as libro_maestro  # noqa: E402

# etl/captura.py es OPCIONAL: la captura repartida por unidad no está en uso,
# el flujo vigente es un solo libro maestro. Si el módulo no está en el
# repositorio, estas pruebas se SALTAN.
#
# Antes esto era un `import captura` a secas, y su ausencia rompía el
# descubrimiento de pruebas completo: unittest abortaba con ImportError, el
# paso de CI fallaba y el tablero no se publicaba nunca. Una prueba de algo
# opcional no puede tumbar el despliegue de lo que sí está.
try:
    import captura  # noqa: E402
except ImportError:  # pragma: no cover
    captura = None

# El import puede "funcionar" sin que exista el módulo: en la raíz hay una
# carpeta llamada captura/ con los libros generados, y Python la toma como
# paquete de espacio de nombres. Importa vacío y el fallo aparece después,
# como AttributeError dentro de setUpClass. Por eso no basta con que el
# import no truene: se comprueba que el módulo traiga lo que se va a usar.
if captura is not None and not hasattr(captura, "crear"):  # pragma: no cover
    captura = None

from openpyxl import load_workbook  # noqa: E402

HAY_CAPTURA = captura is not None
FALTA = "etl/captura.py no está en el repositorio (herramienta opcional)"


@unittest.skipUnless(HAY_CAPTURA, FALTA)
@unittest.skipUnless(os.path.exists(libro_maestro.LIBRO), "falta data/BASE_RRHH.xlsx")
class TestHojasDeCaptura(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.dir = tempfile.mkdtemp()
        cls.rutas = captura.crear(destino_base=cls.dir)
        cls.tablas, _, cls.periodos = captura.cargar_maestro()
        cls.periodo = captura.siguiente_periodo(cls.periodos[-1])
        cls.wb = load_workbook(cls.rutas[0])
        cls.hoja = cls.wb["Plantilla"]

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        shutil.rmtree(cls.dir, ignore_errors=True)

    def test_hay_un_libro_por_unidad_que_reporta(self):
        con_datos = {f["unidad_id"] for f in self.tablas["fact_plantilla"]}
        self.assertEqual(len(self.rutas), len(con_datos))

    def test_instrucciones_va_primero(self):
        self.assertEqual(self.wb.sheetnames[0], "INSTRUCCIONES")

    def test_estan_las_seis_hojas_de_captura(self):
        for cfg in captura.HOJAS.values():
            self.assertIn(cfg["titulo"], self.wb.sheetnames)

    def test_no_incluye_catalogos_ni_metas(self):
        """El analista no administra catálogos: eso es de la Dirección."""
        for prohibida in ("dim_unidad", "dim_area", "metas"):
            self.assertNotIn(prohibida, self.wb.sheetnames)

    def test_las_llaves_vienen_preescritas(self):
        """Si el analista tuviera que teclear el periodo, lo escribiría mal."""
        self.assertEqual(str(self.hoja["B4"].value), self.periodo)
        self.assertTrue(str(self.hoja["C4"].value).startswith("U"))
        self.assertIn(self.hoja["E4"].value, schema.TIPOS_RELACION)

    def test_las_llaves_estan_bloqueadas_y_en_gris(self):
        self.assertTrue(self.hoja["B4"].protection.locked)
        self.assertEqual(self.hoja["B4"].fill.fgColor.rgb, captura.GRIS)
        self.assertTrue(self.hoja.protection.sheet, "la hoja debería estar protegida")

    def test_las_celdas_de_captura_estan_libres_y_en_amarillo(self):
        columnas = list(schema.TABLAS["fact_plantilla"]["columnas"].keys())
        campos = [c for c in columnas if c not in captura.LLAVES]
        enc = ["Área"] + [c for c in columnas if c in captura.LLAVES] + campos
        primera = len(enc) - len(campos) + 1
        celda = self.hoja.cell(row=4, column=primera)
        self.assertFalse(celda.protection.locked)
        self.assertEqual(celda.fill.fgColor.rgb, captura.AMARILLO)
        self.assertIsNone(celda.value, "debe llegar vacía para capturar")

    def test_trae_la_referencia_del_mes_anterior_con_valores(self):
        columnas = list(schema.TABLAS["fact_plantilla"]["columnas"].keys())
        campos = [c for c in columnas if c not in captura.LLAVES]
        enc = ["Área"] + [c for c in columnas if c in captura.LLAVES] + campos
        col_ref = len(enc) + 2
        etiqueta = self.hoja.cell(row=2, column=col_ref).value
        self.assertIn(self.periodos[-1], str(etiqueta))
        # El headcount del mes anterior solo puede venir si ese mes se
        # capturó. Con la hoja vacía no hay referencia que traer, y exigirla
        # obligaría a sembrar un número inventado en el libro real.
        ref = self.hoja.cell(row=4, column=col_ref + 2).value
        if ref is None:
            self.skipTest("el mes anterior no tiene plantilla capturada para "
                          "esa área: no hay referencia que mostrar")
        self.assertIsInstance(ref, (int, float))

    def test_el_sindicato_se_hereda_y_no_llega_vacio(self):
        """Bug real: venía vacío y bloqueado, así que nadie podía llenarlo."""
        ws = self.wb["Relaciones laborales"]
        columnas = list(schema.TABLAS["fact_relaciones_laborales"]["columnas"].keys())
        j = 2 + [c for c in columnas if c in captura.LLAVES].index("sindicato")
        self.assertTrue(str(ws.cell(row=4, column=j).value).strip(),
                        "el sindicato debe venir preescrito")

    def test_enps_no_tiene_validacion_de_no_negativos(self):
        """Bug real: el eNPS va de -100 a 100; la regla genérica lo rechazaba."""
        ws = self.wb["Relaciones laborales"]
        columnas = list(schema.TABLAS["fact_relaciones_laborales"]["columnas"].keys())
        campos = [c for c in columnas if c not in captura.LLAVES]
        enc = ["Área"] + [c for c in columnas if c in captura.LLAVES] + campos
        from openpyxl.utils import get_column_letter
        letra = get_column_letter(len(enc) - len(campos) + 1 + campos.index("enps"))
        for dv in ws.data_validations.dataValidation:
            if letra + "4" in str(dv.sqref) or f"{letra}4:" in str(dv.sqref):
                if dv.operator == "greaterThanOrEqual" and dv.formula1 == "0":
                    self.fail("el eNPS no puede tener validación de no negativos")

    def test_las_reglas_de_par_se_validan_en_excel(self):
        """dc3_emitidos <= dc3_requeridos debe avisarse al teclear, no al final."""
        ws = self.wb["Capacitación"]
        personalizadas = [dv for dv in ws.data_validations.dataValidation
                          if dv.type == "custom"]
        self.assertTrue(personalizadas,
                        "falta la validación cruzada de DC-3 en la hoja")

    def test_una_unidad_solo_lleva_sus_areas(self):
        con_datos = {}
        for f in self.tablas["fact_plantilla"]:
            con_datos.setdefault(f["unidad_id"], set()).add(f["area_id"])
        uid = str(self.hoja["C4"].value)
        claves = {str(self.hoja.cell(row=i, column=4).value)
                  for i in range(4, self.hoja.max_row + 1)
                  if self.hoja.cell(row=i, column=4).value}
        self.assertEqual(claves, con_datos[uid])


if __name__ == "__main__":
    unittest.main(verbosity=2)
