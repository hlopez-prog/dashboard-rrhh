"""
La base de datos del Dashboard de RRHH es un libro de Excel.

    python3 etl/excel.py crear     # genera data/BASE_RRHH.xlsx desde data/csv/
    python3 etl/excel.py leer      # verifica que el libro se lea correctamente

`data/BASE_RRHH.xlsx` es la FUENTE DE VERDAD: es el archivo que RRHH edita cada
mes. Una hoja por tabla, con los mismos encabezados que valida el ETL, listas
desplegables en las columnas de catálogo y hoja de instrucciones.

`data/csv/` se genera automáticamente a partir del libro en cada build. Existe
por una razón concreta: un .xlsx es binario y git no puede mostrar qué cambió
dentro. El export en CSV hace que el diff del cierre mensual sea legible y
auditable en GitHub, línea por línea.

Reglas de diseño del libro:
  - No lleva fórmulas en las hojas de datos. Es un almacén de datos, no un
    modelo: todo el cálculo vive en el tablero (public/assets/js/kpi.js), de
    modo que exista una sola definición de cada KPI.
  - La hoja LEEME sí lleva fórmulas de conteo, para que quien captura vea de
    inmediato cuántos registros tiene cargados en cada hoja.
  - La celda de origen de datos (DEMO / REAL) controla el aviso naranja del
    tablero. Vive en el libro porque es quien captura el que sabe qué cargó.
"""
import csv
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import schema  # noqa: E402

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LIBRO = os.path.join(RAIZ, "data", "BASE_RRHH.xlsx")
CSV_DIR = os.path.join(RAIZ, "data", "csv")

HOJA_LEEME = "LEEME"
CELDA_ORIGEN = "C7"          # DEMO | REAL
CELDA_ORGANIZACION = "C8"
FILA_ENCABEZADO = 1
MAX_FILAS_VALIDACION = 5000  # alcance de las listas desplegables

# --- Paleta institucional -------------------------------------------------
NEGRO = "FF000000"
BLANCO = "FFFFFFFF"
VERDE = "FFD9EAD3"
NARANJA = "FFFCE5CD"
GRIS = "FFF2F2F2"

FUENTE = "Arial"

F_ENCABEZADO = Font(name=FUENTE, bold=True, color=BLANCO, size=10)
R_ENCABEZADO = PatternFill("solid", fgColor=NEGRO)
F_NORMAL = Font(name=FUENTE, size=10)
F_TITULO = Font(name=FUENTE, bold=True, size=14)
F_SUBTITULO = Font(name=FUENTE, bold=True, size=11)
F_ENTRADA = Font(name=FUENTE, bold=True, size=10, color="FF0000FF")  # azul = capturar
BORDE_FINO = Border(*[Side(style="thin", color="FFD5D5D5")] * 4)

FORMATO = {
    "s": "@",
    "p": "@",
    "i": "#,##0",
    "f": "#,##0.00",
}

# Columnas que son montos en MXN: formato de moneda para que no se lean como conteos
COLUMNAS_MXN = {
    "costo_ordinario", "costo_horas_extra", "costo_prestaciones",
    "presupuesto_costo_laboral", "inversion_mxn",
}

# Catálogos cerrados: lista desplegable escrita en la validación
CATALOGOS = {
    "tipo_relacion": schema.TIPOS_RELACION,
    "turno": schema.TURNOS,
    "tipo_area": schema.TIPOS_AREA,
    "direccion": schema.DIRECCION_META,
}

# Columnas que deben existir en una dimensión: lista desplegable por rango
REFERENCIAS = {
    "unidad_id": ("dim_unidad", "A"),
    "area_id": ("dim_area", "A"),
}

ANCHOS = {"s": 26, "p": 11, "i": 13, "f": 15}


# =====================================================================
# Escritura del libro
# =====================================================================

def _leer_csv(tabla):
    ruta = os.path.join(CSV_DIR, f"{tabla}.csv")
    if not os.path.exists(ruta):
        return []
    with open(ruta, newline="", encoding="utf-8-sig") as f:
        lector = csv.reader(f)
        next(lector, None)
        return [fila for fila in lector if any(c.strip() for c in fila)]


def _convertir(tipo, valor):
    v = (valor or "").strip()
    if tipo in ("s", "p"):
        return v
    if v == "":
        return None
    try:
        return int(float(v)) if tipo == "i" else float(v)
    except ValueError:
        return v


def _hoja_leeme(wb, tablas, origen, organizacion):
    ws = wb.create_sheet(HOJA_LEEME, 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 52

    ws["B2"] = "BASE DE DATOS — DASHBOARD DE RRHH"
    ws["B2"].font = F_TITULO
    ws["B3"] = "Minera Rio Tinto · una hoja por tabla · no borrar ni renombrar hojas"
    ws["B3"].font = Font(name=FUENTE, size=10, italic=True)
    for c in ("B2", "C2", "D2", "E2", "B3", "C3", "D3", "E3"):
        ws[c].fill = PatternFill("solid", fgColor=GRIS)

    ws["B5"] = "CONFIGURACIÓN"
    ws["B5"].font = F_SUBTITULO
    ws["B5"].fill = PatternFill("solid", fgColor=GRIS)
    ws["C5"].fill = PatternFill("solid", fgColor=GRIS)
    ws["E5"].fill = PatternFill("solid", fgColor=GRIS)

    ws["B7"] = "Origen de los datos"
    ws["B7"].font = F_NORMAL
    ws[CELDA_ORIGEN] = origen
    ws[CELDA_ORIGEN].font = F_ENTRADA
    ws[CELDA_ORIGEN].fill = PatternFill("solid", fgColor=NARANJA if origen == "DEMO" else VERDE)
    ws["E7"] = ('DEMO enciende el aviso naranja de datos de demostración en el '
                'tablero. Cambia a REAL cuando cargues el cierre de la operación.')
    ws["E7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["E7"].font = F_NORMAL

    ws["B8"] = "Nombre de la organización"
    ws["B8"].font = F_NORMAL
    ws[CELDA_ORGANIZACION] = organizacion
    ws[CELDA_ORGANIZACION].font = F_ENTRADA
    ws[CELDA_ORGANIZACION].fill = PatternFill("solid", fgColor=VERDE)
    ws["E8"] = "Aparece bajo el título del tablero."
    ws["E8"].font = F_NORMAL

    dv_origen = DataValidation(
        type="list", formula1='"DEMO,REAL"', allow_blank=False,
        errorTitle="Valor no permitido", error="Escribe DEMO o REAL.",
        promptTitle="Origen de los datos", prompt="DEMO o REAL",
    )
    ws.add_data_validation(dv_origen)
    dv_origen.add(ws[CELDA_ORIGEN])

    ws["B10"] = "REGISTROS CARGADOS"
    ws["B10"].font = F_SUBTITULO
    ws["B10"].fill = PatternFill("solid", fgColor=GRIS)
    ws["C10"].fill = PatternFill("solid", fgColor=GRIS)
    ws["D10"].fill = PatternFill("solid", fgColor=GRIS)
    ws["E10"].fill = PatternFill("solid", fgColor=GRIS)

    fila = 11
    for c, txt in (("B", "Hoja"), ("C", "Registros"), ("D", "Columnas"),
                   ("E", "Qué contiene")):
        ws[f"{c}{fila}"] = txt
        ws[f"{c}{fila}"].font = F_ENCABEZADO
        ws[f"{c}{fila}"].fill = R_ENCABEZADO
    fila += 1

    QUE_CONTIENE = {
        "dim_unidad": "Catálogo de unidades mineras y corporativo",
        "dim_area": "Catálogo de áreas y si son críticas",
        "fact_plantilla": "Headcount, dotación autorizada y puestos críticos",
        "fact_movimientos": "Altas, bajas, vacantes y días de cobertura",
        "fact_ausentismo": "Horas programadas vs ausencia e incapacidades",
        "fact_nomina": "Costo laboral, horas extra y toneladas movidas",
        "fact_capacitacion": "Plan vs ejecutado, DC-3 y competencias",
        "fact_relaciones_laborales": "Sindicato, conflictos, eNPS y revisión de CCT",
        "metas": "Meta autorizada por indicador",
    }
    primera_conteo = fila
    for tabla in schema.ORDEN_CARGA:
        columnas = list(schema.TABLAS[tabla]["columnas"].keys())
        letra = get_column_letter(len(columnas))
        ws[f"B{fila}"] = tabla
        ws[f"B{fila}"].font = F_NORMAL
        # Fórmula real: el conteo se recalcula al capturar, no queda congelado.
        ws[f"C{fila}"] = f"=COUNTA({tabla}!A{FILA_ENCABEZADO + 1}:A{MAX_FILAS_VALIDACION})"
        ws[f"C{fila}"].font = F_NORMAL
        ws[f"C{fila}"].number_format = "#,##0"
        ws[f"D{fila}"] = f"=COUNTA({tabla}!A{FILA_ENCABEZADO}:{letra}{FILA_ENCABEZADO})"
        ws[f"D{fila}"].font = F_NORMAL
        ws[f"E{fila}"] = QUE_CONTIENE.get(tabla, "")
        ws[f"E{fila}"].font = F_NORMAL
        for c in "BCDE":
            ws[f"{c}{fila}"].border = BORDE_FINO
        fila += 1

    ws[f"B{fila}"] = "TOTAL"
    ws[f"B{fila}"].font = F_SUBTITULO
    ws[f"C{fila}"] = f"=SUM(C{primera_conteo}:C{fila - 1})"
    ws[f"C{fila}"].font = F_SUBTITULO
    ws[f"C{fila}"].number_format = "#,##0"
    for c in "BCDE":
        ws[f"{c}{fila}"].fill = PatternFill("solid", fgColor=GRIS)

    fila += 3
    ws[f"B{fila}"] = "CÓMO CAPTURAR"
    ws[f"B{fila}"].font = F_SUBTITULO
    ws[f"B{fila}"].fill = PatternFill("solid", fgColor=GRIS)
    for c in "CDE":
        ws[f"{c}{fila}"].fill = PatternFill("solid", fgColor=GRIS)
    fila += 2

    INSTRUCCIONES = [
        ("1. Una fila por combinación de llave",
         "Cada hoja tiene una llave primaria. Dos filas con la misma llave hacen "
         "fallar la validación. La llave está anotada en el comentario del primer "
         "encabezado de cada hoja."),
        ("2. El periodo se escribe AAAA-MM",
         "Por ejemplo 2026-08. La columna ya está formateada como texto para que "
         "Excel no la convierta en fecha. No cambies ese formato."),
        ("3. No borrar ni renombrar hojas ni columnas",
         "El ETL valida los encabezados exactos y en el mismo orden. Si necesitas "
         "una columna nueva, se agrega también en etl/schema.py."),
        ("4. Las celdas azules son de captura",
         "En esta hoja, azul significa que tú decides el valor. En las hojas de "
         "datos, todas las celdas son de captura."),
        ("5. Las columnas de catálogo tienen desplegable",
         "tipo_relacion, turno, tipo_area, direccion, unidad_id y area_id validan "
         "contra una lista. Si el desplegable no ofrece lo que necesitas, primero "
         "agrégalo a la hoja de catálogo correspondiente."),
        ("6. Montos en pesos, sin signo ni separador",
         "Escribe 1250000, no $1,250,000. El formato de celda ya muestra los "
         "separadores."),
        ("7. Agregar meses, no sobrescribirlos",
         "El tablero calcula tasas anualizadas con ventana móvil de 12 meses. "
         "Borrar historia rompe esos indicadores. Se agregan filas del mes nuevo."),
        ("8. Al terminar, subir este archivo a GitHub",
         "El repositorio valida y republica el tablero solo. Si la validación "
         "falla, el tablero conserva el cierre anterior en lugar de mostrar un "
         "dato equivocado."),
    ]
    for titulo, detalle in INSTRUCCIONES:
        ws[f"B{fila}"] = titulo
        ws[f"B{fila}"].font = Font(name=FUENTE, bold=True, size=10)
        ws[f"B{fila}"].alignment = Alignment(vertical="top", wrap_text=True)
        ws[f"C{fila}"] = detalle
        ws[f"C{fila}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"C{fila}"].font = F_NORMAL
        ws.merge_cells(f"C{fila}:E{fila}")
        ws.row_dimensions[fila].height = 30
        fila += 1

    fila += 1
    ws[f"B{fila}"] = "Diccionario de datos completo: plantillas/DICCIONARIO.md del repositorio."
    ws[f"B{fila}"].font = Font(name=FUENTE, size=9, italic=True)
    return ws


def _hoja_tabla(wb, tabla, filas):
    definicion = schema.TABLAS[tabla]
    columnas = list(definicion["columnas"].items())
    ws = wb.create_sheet(tabla)

    for j, (col, tipo) in enumerate(columnas, start=1):
        celda = ws.cell(row=FILA_ENCABEZADO, column=j, value=col)
        celda.font = F_ENCABEZADO
        celda.fill = R_ENCABEZADO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = ANCHOS[tipo]

    ws.cell(row=FILA_ENCABEZADO, column=1).comment = None
    ws.freeze_panes = f"A{FILA_ENCABEZADO + 1}"
    ws.auto_filter.ref = (f"A{FILA_ENCABEZADO}:"
                          f"{get_column_letter(len(columnas))}{FILA_ENCABEZADO}")
    ws.row_dimensions[FILA_ENCABEZADO].height = 30

    for i, fila in enumerate(filas, start=FILA_ENCABEZADO + 1):
        for j, (col, tipo) in enumerate(columnas, start=1):
            valor = _convertir(tipo, fila[j - 1] if j - 1 < len(fila) else "")
            celda = ws.cell(row=i, column=j, value=valor)
            celda.font = F_NORMAL
            celda.number_format = "#,##0" if col in COLUMNAS_MXN else FORMATO[tipo]

    # --- Validaciones -------------------------------------------------
    ultima = max(len(filas) + FILA_ENCABEZADO + 1, MAX_FILAS_VALIDACION)
    for j, (col, tipo) in enumerate(columnas, start=1):
        letra = get_column_letter(j)
        rango = f"{letra}{FILA_ENCABEZADO + 1}:{letra}{ultima}"

        if col in CATALOGOS:
            dv = DataValidation(
                type="list", formula1='"' + ",".join(CATALOGOS[col]) + '"',
                allow_blank=True, showErrorMessage=True,
                errorTitle="Valor fuera de catálogo",
                error=f"Valores permitidos: {', '.join(CATALOGOS[col])}",
            )
            ws.add_data_validation(dv)
            dv.add(rango)

        elif col in REFERENCIAS and tabla not in ("dim_unidad", "dim_area"):
            hoja_ref, col_ref = REFERENCIAS[col]
            dv = DataValidation(
                type="list",
                formula1=f"={hoja_ref}!${col_ref}${FILA_ENCABEZADO + 1}:"
                        f"${col_ref}${MAX_FILAS_VALIDACION}",
                allow_blank=True, showErrorMessage=True,
                errorTitle="Clave inexistente",
                error=f"Esa clave no existe en la hoja {hoja_ref}. "
                      f"Agrégala primero al catálogo.",
            )
            ws.add_data_validation(dv)
            dv.add(rango)

        elif tipo == "p":
            dv = DataValidation(
                type="textLength", operator="equal", formula1="7",
                allow_blank=True, showErrorMessage=True,
                errorTitle="Periodo mal escrito",
                error="El periodo se escribe AAAA-MM, por ejemplo 2026-08.",
                promptTitle="Periodo", prompt="AAAA-MM (ejemplo: 2026-08)",
            )
            ws.add_data_validation(dv)
            dv.add(rango)

        elif tipo in ("i", "f") and col not in ("es_corporativo", "es_critica"):
            dv = DataValidation(
                type="decimal", operator="greaterThanOrEqual", formula1="0",
                allow_blank=True, showErrorMessage=True,
                errorTitle="Valor inválido",
                error="Este campo no admite valores negativos.",
            )
            ws.add_data_validation(dv)
            dv.add(rango)

    # La llave primaria queda anotada donde quien captura la va a ver
    pk = ", ".join(definicion["pk"])
    from openpyxl.comments import Comment
    ws.cell(row=FILA_ENCABEZADO, column=1).comment = Comment(
        f"Llave primaria de esta hoja: {pk}.\n"
        f"No puede haber dos filas con la misma combinación.", "ETL")
    return ws


def crear():
    wb = Workbook()
    wb.remove(wb.active)

    origen = os.environ.get("ORIGEN_DATOS", "DEMO").strip().upper()
    if origen not in ("DEMO", "REAL"):
        origen = "DEMO"

    tablas = {t: _leer_csv(t) for t in schema.ORDEN_CARGA}
    _hoja_leeme(wb, tablas, origen, os.environ.get("ORG_NOMBRE", "Minera Rio Tinto"))
    for tabla in schema.ORDEN_CARGA:
        _hoja_tabla(wb, tabla, tablas[tabla])

    wb.properties.title = "Base de datos — Dashboard de RRHH"
    wb.properties.creator = "ETL Dashboard de RRHH"
    os.makedirs(os.path.dirname(LIBRO), exist_ok=True)
    wb.save(LIBRO)

    total = sum(len(v) for v in tablas.values())
    kb = os.path.getsize(LIBRO) / 1024
    print(f"  data/BASE_RRHH.xlsx  ({kb:.0f} KB · {len(tablas)} hojas · "
          f"{total:,} registros · origen {origen})")
    return LIBRO


# =====================================================================
# Lectura del libro
# =====================================================================

def leer(ruta=LIBRO):
    """
    Devuelve (tablas, config).

    tablas: {nombre_tabla: [ {columna: valor}, ... ]} con '_n' = fila de Excel,
            para que los mensajes de error apunten a la celda real.
    config: {'origen': 'DEMO'|'REAL', 'organizacion': str}
    """
    wb = load_workbook(ruta, data_only=True, read_only=False)
    faltantes = [t for t in schema.ORDEN_CARGA if t not in wb.sheetnames]
    if faltantes:
        raise SystemExit(
            f"El libro {os.path.basename(ruta)} no tiene las hojas: "
            f"{', '.join(faltantes)}.\n"
            f"Hojas encontradas: {', '.join(wb.sheetnames)}")

    config = {"origen": "REAL", "organizacion": "Minera Rio Tinto"}
    if HOJA_LEEME in wb.sheetnames:
        ws = wb[HOJA_LEEME]
        origen = ws[CELDA_ORIGEN].value
        if isinstance(origen, str) and origen.strip().upper() in ("DEMO", "REAL"):
            config["origen"] = origen.strip().upper()
        org = ws[CELDA_ORGANIZACION].value
        if isinstance(org, str) and org.strip():
            config["organizacion"] = org.strip()

    tablas = {}
    for tabla in schema.ORDEN_CARGA:
        definicion = schema.TABLAS[tabla]["columnas"]
        esperados = list(definicion.keys())
        ws = wb[tabla]

        encabezados = []
        for j in range(1, len(esperados) + 1):
            v = ws.cell(row=FILA_ENCABEZADO, column=j).value
            encabezados.append(str(v).strip() if v is not None else "")
        if encabezados != esperados:
            raise SystemExit(
                f"Hoja '{tabla}': encabezados incorrectos.\n"
                f"  esperado: {esperados}\n"
                f"  recibido: {encabezados}\n"
                f"No renombres ni reordenes columnas del libro.")

        filas = []
        for i in range(FILA_ENCABEZADO + 1, ws.max_row + 1):
            valores = [ws.cell(row=i, column=j).value
                       for j in range(1, len(esperados) + 1)]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in valores):
                continue
            fila = {}
            for col, v in zip(esperados, valores):
                tipo = definicion[col]
                if v is None:
                    fila[col] = "" if tipo in ("s", "p") else ""
                elif tipo in ("s", "p"):
                    # Excel puede devolver un número o una fecha donde se espera
                    # texto: se normaliza aquí en lugar de fallar más adelante.
                    if hasattr(v, "strftime"):
                        fila[col] = v.strftime("%Y-%m")
                    elif isinstance(v, float) and v.is_integer():
                        fila[col] = str(int(v))
                    else:
                        fila[col] = str(v).strip()
                else:
                    fila[col] = v
            fila["_n"] = i
            filas.append(fila)
        tablas[tabla] = filas

    wb.close()
    return tablas, config


def exportar_csv(tablas):
    """Export legible para que el diff mensual sea auditable en git."""
    os.makedirs(CSV_DIR, exist_ok=True)
    for tabla, filas in tablas.items():
        columnas = list(schema.TABLAS[tabla]["columnas"].keys())
        ruta = os.path.join(CSV_DIR, f"{tabla}.csv")
        with open(ruta, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(columnas)
            for fila in filas:
                w.writerow([fila.get(c, "") for c in columnas])
    print(f"  data/csv/  ({len(tablas)} archivos exportados para diff en git)")


if __name__ == "__main__":
    accion = sys.argv[1] if len(sys.argv) > 1 else "crear"
    if accion == "crear":
        crear()
    elif accion == "leer":
        tablas, config = leer()
        print(f"  origen: {config['origen']} · organización: {config['organizacion']}")
        for t, f in tablas.items():
            print(f"  {t:<28} {len(f):>6} registros")
    else:
        raise SystemExit("Uso: python3 etl/excel.py [crear|leer]")
