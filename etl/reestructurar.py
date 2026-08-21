"""
Reestructura la base cuando cambia el catálogo de unidades o de áreas.

    python3 etl/reestructurar.py <libro_con_catalogos_nuevos.xlsx>

El problema que resuelve: cuando alguien actualiza `dim_unidad` o `dim_area`
—porque abrió una unidad, cerró otra o reorganizó áreas— las tablas de hechos
quedan apuntando a la estructura vieja. Las claves siguen existiendo, así que la
validación referencial pasa, pero los números ya significan otra cosa: la fila
de U01/A01 que era "Mina La Esperanza / Mina Subterránea" pasa a leerse como
"Unidad Cieneguita / Planta de Beneficio". Y las unidades nuevas quedan sin
datos.

Lo que hace este script:
  1. Toma los catálogos del libro nuevo y los limpia (espacios, celdas vacías).
  2. Deriva qué áreas existen en cada unidad a partir de su `tipo_operacion`.
  3. Regenera 24 meses de hechos sintéticos coherentes con esa estructura.
  4. Conserva la hoja `metas` tal como venía.

IMPORTANTE: los números que genera son SINTÉTICOS. Sirven para que el tablero
funcione con los nombres reales de la operación mientras RRHH captura las cifras
del cierre. El origen se deja en DEMO para que el aviso naranja siga encendido.
"""
import csv
import math
import os
import random
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import schema  # noqa: E402

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_DIR = os.path.join(RAIZ, "data", "csv")

random.seed(20260821)

MES_FINAL = (2026, 7)
N_MESES = 24

# --- Qué áreas existen en cada tipo de operación -------------------------
# Se expresa por TIPO de área, no por clave, para que siga funcionando si
# mañana cambian los area_id.
MEZCLA = {
    "Planta":        ["Planta", "Metalurgia", "Mantenimiento", "Seguridad", "Administración"],
    "Cielo abierto": ["Mina", "Mantenimiento", "Seguridad", "Administración"],
    "Subterránea":   ["Mina", "Mantenimiento", "Seguridad", "Administración"],
    "Exploración":   ["Exploración", "Mantenimiento", "Seguridad", "Administración"],
    "Corporativo":   ["Seguridad", "Administración"],
}
MEZCLA_POR_DEFECTO = ["Mantenimiento", "Seguridad", "Administración"]

# --- Tamaño típico de plantilla propia por tipo de área ------------------
# SUPUESTO a ajustar con las cifras reales de la operación.
BASE_PROPIO = {
    "Mina": 420, "Planta": 255, "Metalurgia": 95, "Exploración": 70,
    "Mantenimiento": 165, "Seguridad": 28, "Administración": 42,
}
# Contratistas como proporción de la plantilla propia del área
PROP_CONTRATISTA = {
    "Mina": 0.72, "Planta": 0.28, "Metalurgia": 0.18, "Exploración": 0.55,
    "Mantenimiento": 0.60, "Seguridad": 0.15, "Administración": 0.14,
}
TURNO = {
    "Mina": "14x14", "Planta": "4x3", "Metalurgia": "4x3", "Exploración": "7x7",
    "Mantenimiento": "4x3", "Seguridad": "Administrativo",
    "Administración": "Administrativo",
}
SALARIO_MES = {
    "Mina": 26800, "Planta": 22400, "Metalurgia": 28900, "Exploración": 31200,
    "Mantenimiento": 25600, "Seguridad": 27300, "Administración": 34500,
}
# Áreas con producción atribuible, y toneladas por persona al mes
TONELADAS_POR_PERSONA = {"Mina": 1180, "Planta": 1520, "Metalurgia": 240}

# Perfiles de desempeño, para que el semáforo del tablero muestre una mezcla
# realista en lugar de todo verde o todo rojo.
PERFILES = [
    {"rot": 1.45, "aus": 1.28, "he": 1.42, "cap": 0.955, "riesgo": 4},
    {"rot": 1.00, "aus": 0.92, "he": 1.00, "cap": 0.995, "riesgo": 2},
    {"rot": 1.55, "aus": 1.12, "he": 1.28, "cap": 0.970, "riesgo": 3},
    {"rot": 0.78, "aus": 0.80, "he": 0.82, "cap": 1.020, "riesgo": 2},
    {"rot": 1.20, "aus": 1.05, "he": 1.15, "cap": 0.985, "riesgo": 3},
    {"rot": 0.90, "aus": 0.88, "he": 0.95, "cap": 1.005, "riesgo": 2},
    {"rot": 1.10, "aus": 0.95, "he": 0.90, "cap": 1.000, "riesgo": 2},
]
PERFIL_CORPORATIVO = {"rot": 0.78, "aus": 0.62, "he": 0.55, "cap": 1.030, "riesgo": 1}
# Factor de tamaño por unidad, para que no salgan todas iguales
FACTOR_TAMANO = [1.00, 1.15, 1.05, 0.80, 0.62, 0.90, 1.00, 1.00, 0.95, 0.85]


def periodos():
    y, m = MES_FINAL
    out = []
    for _ in range(N_MESES):
        out.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return list(reversed(out))


PERIODOS = periodos()


def estacional(i, amp, fase=0.0):
    return amp * math.sin(2 * math.pi * ((i % 12) / 12.0) + fase)


def tendencia(i, delta):
    return delta * (i / (N_MESES - 1))


def r(v, pct=0.06):
    return v * (1 + random.uniform(-pct, pct))


def texto(v, por_defecto=""):
    if v is None:
        return por_defecto
    return str(v).strip()


def entero(v, por_defecto=0):
    if v is None or str(v).strip() == "":
        return por_defecto
    try:
        return int(float(v))
    except ValueError:
        return por_defecto


def leer_catalogos(ruta):
    """Lee y limpia dim_unidad, dim_area y metas del libro dado."""
    wb = load_workbook(ruta, data_only=True)
    avisos = []

    unidades = []
    ws = wb["dim_unidad"]
    for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not fila or not texto(fila[0]):
            continue
        uid = texto(fila[0])
        nombre = texto(fila[1], uid)
        tipo_op = texto(fila[3], "Planta")
        mineral = texto(fila[4])
        corp_bruto = fila[5]
        corp = entero(corp_bruto, 1 if tipo_op == "Corporativo" else 0)
        if corp_bruto is None or str(corp_bruto).strip() == "":
            avisos.append(f"dim_unidad fila {i} ({nombre}): 'es_corporativo' venía "
                          f"vacía, se puso {corp}")
        if not mineral:
            mineral = "N/A" if tipo_op == "Corporativo" else "Oro-Plata"
            avisos.append(f"dim_unidad fila {i} ({nombre}): 'mineral_principal' "
                          f"venía vacía, se puso '{mineral}'")
        unidades.append({
            "unidad_id": uid, "unidad": nombre, "estado": texto(fila[2], "Chihuahua"),
            "tipo_operacion": tipo_op, "mineral_principal": mineral,
            "es_corporativo": corp,
        })

    areas = []
    ws = wb["dim_area"]
    for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not fila or not texto(fila[0]):
            continue
        aid = texto(fila[0])
        nombre = texto(fila[1], aid)
        tipo = texto(fila[2])
        if tipo not in schema.TIPOS_AREA:
            avisos.append(f"dim_area fila {i} ({nombre}): 'tipo_area' = '{tipo}' "
                          f"no está en el catálogo; revísalo")
        crit_bruto = fila[3]
        # Por defecto, crítica salvo las áreas de soporte
        crit = entero(crit_bruto, 0 if tipo in ("Seguridad", "Administración") else 1)
        if crit_bruto is None or str(crit_bruto).strip() == "":
            avisos.append(f"dim_area fila {i} ({nombre}): 'es_critica' venía vacía, "
                          f"se puso {crit}")
        areas.append({"area_id": aid, "area": nombre, "tipo_area": tipo,
                      "es_critica": crit})

    metas = []
    ws = wb["metas"]
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila or not texto(fila[0]):
            continue
        metas.append({
            "kpi": texto(fila[0]), "nombre": texto(fila[1]),
            "meta": float(fila[2]) if fila[2] is not None else 0.0,
            "direccion": texto(fila[3], "menor_mejor"),
            "unidad_medida": texto(fila[4]),
        })
    wb.close()
    return unidades, areas, metas, avisos


def construir_estructura(unidades, areas):
    """unidad_id -> [area, ...] según el tipo de operación."""
    por_tipo = {}
    for a in areas:
        por_tipo.setdefault(a["tipo_area"], []).append(a)

    estructura = {}
    for u in unidades:
        tipos = MEZCLA.get(u["tipo_operacion"], MEZCLA_POR_DEFECTO)
        elegidas = []
        for t in tipos:
            elegidas.extend(por_tipo.get(t, []))
        if not elegidas:                      # ningún área del tipo esperado
            elegidas = areas[:]
        estructura[u["unidad_id"]] = elegidas
    return estructura


def generar(unidades, areas, metas):
    estructura = construir_estructura(unidades, areas)
    perfil = {}
    factor = {}
    for k, u in enumerate(unidades):
        uid = u["unidad_id"]
        perfil[uid] = (PERFIL_CORPORATIVO if u["es_corporativo"]
                       else PERFILES[k % len(PERFILES)])
        factor[uid] = (1.0 if u["es_corporativo"]
                       else FACTOR_TAMANO[k % len(FACTOR_TAMANO)])

    plantilla, movs, aus, nom, cap, rl = [], [], [], [], [], []

    for i, per in enumerate(PERIODOS):
        for u in unidades:
            uid = u["unidad_id"]
            p = perfil[uid]
            f = factor[uid]
            hc_unidad = 0

            for a in estructura[uid]:
                aid, tipo = a["area_id"], a["tipo_area"]
                critica = a["es_critica"]

                if u["es_corporativo"]:
                    base = 12 if tipo == "Seguridad" else 96
                else:
                    base = BASE_PROPIO.get(tipo, 40) * f

                crec = 1 + tendencia(i, 0.06) + estacional(i, 0.015)
                hc_prop = max(1, round(r(base * crec, 0.03)))
                hc_cont = max(0, round(r(base * PROP_CONTRATISTA.get(tipo, 0.2)
                                         * (1 + tendencia(i, 0.11)), 0.07)))
                hc_unidad += hc_prop

                autorizada = int(base * 1.06)
                pct_mujeres = (0.30 if tipo == "Administración"
                               else 0.19 if tipo == "Seguridad" else 0.09)
                antig = r(78 - 26 * (p["rot"] - 0.8), 0.08)
                turno = TURNO.get(tipo, "Administrativo")
                if tipo == "Mina" and u["tipo_operacion"] == "Subterránea":
                    turno = "7x7"

                pc_tot = max(2, int(hc_prop * (0.16 if critica else 0.07)))
                brecha = 0.042 * p["rot"] + max(0.0, estacional(i, 0.02))
                pc_ok = max(0, int(pc_tot * (1 - brecha)))

                plantilla.append([per, uid, aid, "Propio", turno, hc_prop,
                                  autorizada, int(hc_prop * pct_mujeres),
                                  round(antig, 1), pc_tot, pc_ok])
                if hc_cont > 0:
                    plantilla.append([per, uid, aid, "Contratista", turno, hc_cont,
                                      0, int(hc_cont * pct_mujeres * 0.55),
                                      round(antig * 0.32, 1), 0, 0])

                # Movimientos
                tasa = (0.0092 * p["rot"]) * (1 + estacional(i, 0.22, 1.1)) \
                    * (1 + tendencia(i, 0.10))
                bajas = max(0, round(r(hc_prop * tasa, 0.20)))
                b_vol = int(bajas * random.uniform(0.62, 0.80))
                b_inv = bajas - b_vol
                temprana = int(b_vol * random.uniform(0.22, 0.42))
                altas = max(0, round(r(bajas * random.uniform(0.85, 1.30), 0.18)))
                vac = max(0, round(r(max(0, autorizada - hc_prop) * 0.55
                                     + bajas * 0.7, 0.25)))
                dias_cob = r((32 if critica else 23) * (1 + tendencia(i, 0.14)), 0.12)
                movs.append([per, uid, aid, altas, b_vol, b_inv, temprana, vac,
                             round(dias_cob, 1)])

                # Ausentismo
                hrs_prog = hc_prop * (192 if turno != "Administrativo" else 176)
                tasa_aus = 0.0270 * p["aus"] * (1 + estacional(i, 0.26, 0.4)) \
                    * (1 + tendencia(i, 0.08))
                hrs_aus = min(hrs_prog * 0.20, r(hrs_prog * tasa_aus, 0.14))
                casos = max(0, round(r(hc_prop * 0.026 * p["aus"], 0.22)))
                aus.append([per, uid, aid, round(hrs_prog, 1), round(hrs_aus, 1),
                            casos, int(casos * random.uniform(3.4, 8.2))])

                # Nómina y productividad
                sal = SALARIO_MES.get(tipo, 25000) * (1 + tendencia(i, 0.083))
                costo_ord = r(hc_prop * sal, 0.03)
                tasa_he = 0.0455 * p["he"] * (1 + estacional(i, 0.30, 0.9)) \
                    * (1 + tendencia(i, 0.14))
                hrs_ord = hc_prop * (192 if turno != "Administrativo" else 176)
                hrs_he = r(hrs_ord * tasa_he, 0.16)
                costo_he = hrs_he * (sal / 192) * 2.0
                prest = costo_ord * random.uniform(0.33, 0.39)
                presup = (costo_ord + prest) * 1.045 + costo_he * 0.72
                tpp = TONELADAS_POR_PERSONA.get(tipo, 0)
                ton = r(hc_prop * tpp * (1 + tendencia(i, 0.045)), 0.07) if tpp else 0.0
                nom.append([per, uid, aid, round(costo_ord), round(costo_he),
                            round(prest), round(hrs_ord, 1), round(hrs_he, 1),
                            round(presup), round(ton, 1)])

                # Capacitación
                h_plan = hc_prop * (2.6 if critica else 1.7)
                h_real = r(h_plan * min(1.02, p["cap"] * (1 + estacional(i, 0.10, 2.0))), 0.10)
                dc3_req = max(1, int(hc_prop * 0.30))
                dc3_emi = int(dc3_req * min(0.995, p["cap"] * random.uniform(0.975, 1.015)))
                comp_ok = max(0, int(pc_tot * min(0.985, p["cap"] * random.uniform(0.955, 1.0))))
                cap.append([per, uid, aid, round(h_plan, 1), round(h_real, 1),
                            max(1, int(hc_prop * 0.45)), dc3_req, dc3_emi,
                            round(h_real * random.uniform(255, 420)), pc_tot, comp_ok])

            # Relaciones laborales, una fila por unidad
            if u["es_corporativo"]:
                sindicato, sindicalizados, dias_cct = "No sindicalizado", 0, 0
            else:
                sindicato = "Por confirmar — Relaciones Laborales"
                sindicalizados = int(hc_unidad * random.uniform(0.68, 0.86))
                dias_cct = 365 - ((i * 30 + abs(hash(uid)) % 365) % 365)
            abiertos = max(0, round(r(p["riesgo"] * 1.6 * (1 + tendencia(i, 0.28)), 0.30)))
            rl.append([
                per, uid, sindicato, sindicalizados,
                1 if (not u["es_corporativo"] and random.random() < 0.055 * p["riesgo"]) else 0,
                abiertos, max(0, int(abiertos * random.uniform(0.35, 0.85))),
                dias_cct, p["riesgo"],
                round(r(43 - 6.5 * (p["riesgo"] - 1) + tendencia(i, -0.14) * 30, 0.09), 1),
                round(r(0.72, 0.10), 3),
            ])

    return {
        "dim_unidad": [[u[c] for c in schema.TABLAS["dim_unidad"]["columnas"]]
                       for u in unidades],
        "dim_area": [[a[c] for c in schema.TABLAS["dim_area"]["columnas"]]
                     for a in areas],
        "fact_plantilla": plantilla,
        "fact_movimientos": movs,
        "fact_ausentismo": aus,
        "fact_nomina": nom,
        "fact_capacitacion": cap,
        "fact_relaciones_laborales": rl,
        "metas": [[m[c] for c in schema.TABLAS["metas"]["columnas"]] for m in metas],
    }


def escribir_csv(datos):
    os.makedirs(CSV_DIR, exist_ok=True)
    for tabla in schema.ORDEN_CARGA:
        columnas = list(schema.TABLAS[tabla]["columnas"].keys())
        with open(os.path.join(CSV_DIR, f"{tabla}.csv"), "w", newline="",
                  encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(columnas)
            w.writerows(datos[tabla])
        print(f"  {tabla:<28} {len(datos[tabla]):>6} filas")


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Uso: python3 etl/reestructurar.py <libro.xlsx>")
    ruta = sys.argv[1]
    if not os.path.exists(ruta):
        raise SystemExit(f"No existe: {ruta}")

    print(f"Leyendo catálogos de {os.path.basename(ruta)}\n")
    unidades, areas, metas, avisos = leer_catalogos(ruta)

    print(f"Unidades: {len(unidades)}")
    for u in unidades:
        print(f"  {u['unidad_id']}  {u['unidad']:<26} {u['tipo_operacion']:<15}"
              f"{'corporativo' if u['es_corporativo'] else ''}")
    print(f"\nÁreas: {len(areas)}")
    for a in areas:
        print(f"  {a['area_id']}  {a['area']:<26} {a['tipo_area']:<15}"
              f"{'crítica' if a['es_critica'] else ''}")

    if avisos:
        print(f"\nSe corrigieron {len(avisos)} huecos en los catálogos:")
        for m in avisos:
            print(f"  · {m}")

    estructura = construir_estructura(unidades, areas)
    print("\nÁreas asignadas a cada unidad, según su tipo de operación:")
    for u in unidades:
        nombres = ", ".join(a["area"] for a in estructura[u["unidad_id"]])
        print(f"  {u['unidad']}: {nombres}")

    print(f"\nGenerando {N_MESES} meses ({PERIODOS[0]} → {PERIODOS[-1]})\n")
    datos = generar(unidades, areas, metas)
    escribir_csv(datos)

    por_mes = sum(len([f for f in datos[t] if f[0] == PERIODOS[-1]])
                  for t in schema.ORDEN_CARGA
                  if "periodo" in schema.TABLAS[t]["columnas"])
    print(f"\nFilas por cierre mensual con esta estructura: {por_mes}")
    print("\nSiguiente paso:  python3 etl/excel.py crear && python3 etl/build.py")


if __name__ == "__main__":
    main()
